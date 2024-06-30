import openpyxl
from openpyxl.styles import Font, PatternFill

# Load the workbook and select a sheet
wb = openpyxl.load_workbook('sample.xlsx')
sheet = wb.active

# Reading cell values
print("Reading data from Excel file:")
for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3, values_only=True):
    print(row)

# Write data to the sheet
data = [
    ["Name", "Age", "City"],
    ["Alice", 30, "New York"],
    ["Bob", 25, "Los Angeles"],
    ["Charlie", 35, "Chicago"]
]

for row in data:
    sheet.append(row)

# Apply formatting to header row
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

for cell in sheet["1:1"]:
    cell.font = header_font
    cell.fill = header_fill

# Calculate the average age and write it to the sheet
ages = [row[1] for row in sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=2, values_only=True)]
average_age = sum(ages) / len(ages)
sheet["E1"] = "Average Age"
sheet["E2"] = average_age

# Save the workbook
wb.save('sample.xlsx')
print("Data written and formatted successfully.")
